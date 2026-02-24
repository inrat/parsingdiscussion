"""
Ylis Data Extraction Tool

Collects reply relationship data from forum threads
for academic research purposes.

Rewrite of the original Java implementation (2022):
https://github.com/inrat/parsingdiscussion
"""

__author__ = "Inka Ratia"
__version__ = "2.0"
__license__ = "MIT"

import re
import json
import tkinter as tk
from tkinter import messagebox, filedialog
from datetime import datetime
from bs4 import BeautifulSoup           # HTML parsing library
from playwright.sync_api import sync_playwright  # Browser automation (bypasses bot protection)
from openpyxl import Workbook           # Excel file creation
from openpyxl.styles import Font, PatternFill, Alignment


# ── Data Extractor ────────────────────────────────────────────────────────────────────

def get_csrf_token(html: str) -> str:
    """
    Extract the CSRF (Cross-Site Request Forgery) token from the page HTML.

    Frontend JS is initialized with:
        new App('some_url', 'csrf_token_hex_string')
    We use a regex to capture that hex token, which is needed to authenticate
    our API requests so the server thinks they come from a real browser session.
    """
    match = re.search(r"new App\('[^']+',\s*'([a-f0-9]+)'", html)
    if not match:
        raise ValueError("CSRF token not found in page")
    return match.group(1)


def parse_post(html: str) -> dict:
    """
    Parse a single post's HTML and extract:
      - post_id:  the unique numeric ID of this post
      - user_id:  the anonymous user number within the thread (0 = OP, 1, 2, ...)
      - refs_to:  list of post IDs that this post REPLIES TO

    Reply detection works by looking inside the post's message body for:
      1. <a class="ref" data-post-id="..."> — inline reply links that appear
         when a user clicks "reply" on another post (shows as a small arrow icon)
      2. <div class="post-ref" data-post-id="..."> — quoted reply blocks that
         appear when a post embeds/quotes another post's content

    Both types give us a data-post-id attribute pointing to the target post.
    A single post can contain multiple refs (replying to several posts at once).
    """
    soup = BeautifulSoup(html, "html.parser")

    # Find the root <div class="post"> element
    div = soup.find("div", class_="post")
    if not div:
        return None

    # data-post-id and data-user-id are attributes on the post div
    post_id = int(div["data-post-id"])
    user_id = int(div["data-user-id"])

    # Collect all outgoing references (posts this post replies to)
    refs_to = []
    msg_div = div.find("div", class_="post-message")
    if msg_div:
        # Type 1: inline reply links — <a class="ref" data-post-id="304366258">
        for a in msg_div.find_all("a", class_="ref"):
            if a.get("data-post-id"):
                refs_to.append(int(a["data-post-id"]))

        # Type 2: quote blocks — <div class="post-ref" data-post-id="304366269">
        for d in msg_div.find_all("div", class_="post-ref"):
            if d.get("data-post-id"):
                refs_to.append(int(d["data-post-id"]))

    return {
        "post_id":  post_id,
        "user_id":  user_id,
        "refs_to":  refs_to,  # post IDs that this post replies TO
    }


def fetch_new_replies(page, csrf, thread_id, replies_from_id, visible_ids):
    """
    Call internal API to load additional posts beyond the initial page.

    Threads with many posts are paginated — the first page load only contains
    some posts. This function uses the browser's fetch() (via Playwright's
    page.evaluate) to request more posts from the server.

    Parameters:
      - page:             Playwright page object (the open browser tab)
      - csrf:             CSRF token for authentication
      - thread_id:        numeric thread ID
      - replies_from_id:  "give me posts newer than this post ID"
      - visible_ids:      list of post IDs we already have (so server can skip them)

    Returns the raw JSON response text containing new post HTML fragments.
    """
    result = page.evaluate("""
        async ([csrf, thread_id, replies_from_id, visible_replies]) => {
            const form = new FormData();
            form.append("thread_id", thread_id);
            form.append("replies_from_id", replies_from_id);
            form.append("visible_replies", visible_replies);
            const response = await fetch("/api/community/thread/new-replies", {
                method: "POST",
                headers: { "x-csrf-token": csrf },
                body: form,
            });
            return await response.text();
        }
    """, [csrf, str(thread_id), str(replies_from_id), ",".join(map(str, visible_ids))])
    return result


def get_thread(url: str, status_callback=None) -> list[dict]:
    """
    Main data extraction function. Opens the thread URL in a headless Firefox browser,
    extracts all posts from the initial page, then repeatedly calls the API
    to fetch any remaining posts until the thread is fully loaded.

    Returns a list of post dicts (see parse_post for structure).
    """
    def log(msg):
        if status_callback:
            status_callback(msg)

    # Validate URL and extract board name + thread slug
    # Example: https://ylilauta.org/satunnainen/12345 -> board="satunnainen", slug="12345"
    match = re.search(r"ylilauta\.org/([^/]+)/([^/?#]+)", url)
    if not match:
        raise ValueError("Invalid URL. Expected format: https://ylilauta.org/board/threadid")

    board, slug = match.group(1), match.group(2)
    full_url = f"https://ylilauta.org/{board}/{slug}"

    # Launch a headless Firefox browser via Playwright.
    with sync_playwright() as p:
        try:
            browser = p.firefox.launch(headless=True)
            context = browser.new_context()
            page = context.new_page()

            # Load the thread page and wait for all network activity to settle
            log("Loading page...")
            page.goto(full_url, wait_until="networkidle", timeout=30000)
            html = page.content()

            # Extract CSRF token from the page's JavaScript
            csrf = get_csrf_token(html)
            log("CSRF token found.")

            # Find the thread ID from the page HTML (needed for API calls)
            soup = BeautifulSoup(html, "html.parser")
            thread_elem = soup.find(attrs={"data-thread-id": True})
            if not thread_elem:
                raise ValueError("Could not find thread ID on page. Is the URL correct?")
            thread_id = int(thread_elem["data-thread-id"])
            log(f"Thread ID: {thread_id}")

            all_posts = []     # Accumulates all parsed post dicts
            visible_ids = []   # Tracks post IDs we've already seen (for API pagination)

            # Parse all posts from the initial page load
            for post_div in soup.find_all("div", class_="post"):
                post = parse_post(str(post_div))
                if post:
                    all_posts.append(post)
                    visible_ids.append(post["post_id"])

            log(f"Found {len(all_posts)} posts in initial page load.")

            # Start pagination from the newest post we've seen
            replies_from_id = max(visible_ids) if visible_ids else 0

            # Keep fetching until the API returns no more posts
            while True:
                raw = fetch_new_replies(page, csrf, thread_id, replies_from_id, visible_ids)
                try:
                    data = json.loads(raw)
                except json.JSONDecodeError:
                    raise ValueError(f"Unexpected API response: {raw[:100]}")

                # Empty response = no more posts to load
                if not data:
                    break

                # Parse each new post HTML fragment from the API response
                for post_html in data["posts"]:
                    post = parse_post(post_html)
                    if post:
                        all_posts.append(post)
                        visible_ids.append(post["post_id"])

                # Move the pagination cursor forward
                replies_from_id = max(data["ids"])
                log(f"Fetched more posts... total: {len(all_posts)}")

            browser.close()

        except Exception as e:
            try:
                browser.close()
            except Exception:
                pass
            raise e

    return all_posts


# ── Excel export ───────────────────────────────────────────────────────────────

def save_xlsx(posts: list[dict], filepath: str):
    """
    Export reply relationships to an Excel file with two columns:
      Column A: Poster User ID    (the user who wrote the post)
      Column B: Replied-To User ID (the user they replied to, or #N/A if none)

    Each row represents one reply relationship, NOT one post.
    This means:
      - A post with no replies  -> 1 row:  (user_id, #N/A)
      - A post replying to 1    -> 1 row:  (user_id, target_user_id)
      - A post replying to 3    -> 3 rows: one per target
      - Same user replying to same user in 2 separate posts -> 2 rows
    """

    # Build a lookup table: post_id -> user_id
    # We need this to convert "post X replies to post Y" into
    # "user A replies to user B" (the post IDs are just intermediaries)
    id_to_user = {p["post_id"]: p["user_id"] for p in posts}

    # Build the output rows as a list of (poster, replied_to) tuples.
    # We iterate every post and expand its refs_to list into individual rows.
    rows = []
    for post in posts:
        if not post["refs_to"]:
            # Post doesn't reply to anyone -> mark as #N/A
            rows.append((post["user_id"], "#N/A"))
        else:
            # One row per reply target.
            # Example: user 6 replies to post 100 (user 1) and post 200 (user 2)
            #   -> two rows: (6, 1) and (6, 2)
            for ref_post_id in post["refs_to"]:
                # Look up which user owns the target post.
                # If the target post was deleted or not in our data, fall back to #N/A.
                target_user = id_to_user.get(ref_post_id, "#N/A")
                rows.append((post["user_id"], target_user))

    wb = Workbook()
    ws = wb.active
    ws.title = "Ylis Data"

    # Write data rows directly (no header row — first cell is already a value)
    row_font = Font(name="Arial", size=10)
    for row_idx, (poster_id, replied_to_id) in enumerate(rows, start=1):
        cell_a = ws.cell(row=row_idx, column=1, value=poster_id)
        cell_a.font = row_font
        cell_b = ws.cell(row=row_idx, column=2, value=replied_to_id)
        cell_b.font = row_font

    # Set column widths for readability
    ws.column_dimensions["A"].width = 18
    ws.column_dimensions["B"].width = 20

    wb.save(filepath)


# ── GUI ────────────────────────────────────────────────────────────────────────

class App(tk.Tk):
    """
    Simple tkinter GUI with:
      - A text field for the thread URL
      - A text field + browse button for the output file path
      - A "Get" button that runs the data gatherer and saves the Excel file
      - A status label at the bottom showing progress
    """

    def __init__(self):
        super().__init__()
        self.title("Ylis Data Extraction")
        self.resizable(False, False)
        self.configure(padx=24, pady=20)

        # Title label
        tk.Label(self, text="Ylis Data Exctraction",
                 font=("Arial", 14, "bold")).grid(row=0, column=0, columnspan=2, pady=(0, 15))

        # URL input
        tk.Label(self, text="Thread URL:", font=("Arial", 10)).grid(
            row=1, column=0, columnspan=2, sticky="w")
        self.url_entry = tk.Entry(self, width=55, font=("Arial", 10))
        self.url_entry.insert(0, "https://ylilauta.org/satunnainen/xxxxx")
        self.url_entry.grid(row=2, column=0, columnspan=2, pady=(2, 14))

        # Output file path input + browse button
        tk.Label(self, text="Save as:", font=("Arial", 10)).grid(
            row=3, column=0, sticky="w")
        self.filename_entry = tk.Entry(self, width=42, font=("Arial", 10))
        self.filename_entry.insert(0, "output.xlsx")
        self.filename_entry.grid(row=4, column=0, sticky="w", pady=(2, 14))

        browse_btn = tk.Button(self, text="Browse…", font=("Arial", 10),
                               command=self.browse_save)
        browse_btn.grid(row=4, column=1, sticky="w", padx=(6, 0), pady=(2, 14))

        # Get button
        self.get_btn = tk.Button(self, text="Get", font=("Arial", 11, "bold"),
                                    bg="#2E75B6", fg="white", padx=12, pady=6,
                                    command=self.run_get)
        self.get_btn.grid(row=5, column=0, columnspan=2, pady=(0, 12))

        # Status label (shows progress messages during data extraction)
        self.status_var = tk.StringVar(value="Ready.")
        tk.Label(self, textvariable=self.status_var, font=("Arial", 9),
                 fg="gray", wraplength=400).grid(row=6, column=0, columnspan=2)

    def browse_save(self):
        """Open a file dialog for choosing the output .xlsx path."""
        path = filedialog.asksaveasfilename(
            defaultextension=".xlsx",
            filetypes=[("Excel files", "*.xlsx")],
            initialfile="output.xlsx",
        )
        if path:
            self.filename_entry.delete(0, tk.END)
            self.filename_entry.insert(0, path)

    def set_status(self, msg):
        """Update the status label and force the GUI to redraw immediately."""
        self.status_var.set(msg)
        self.update_idletasks()

    def run_get(self):
        """
        Called when the user clicks "Get". Validates inputs, runs the
        extractor, saves the Excel file, and shows a success/error message.
        The button is disabled during extraction to prevent double-clicks.
        """
        url = self.url_entry.get().strip()
        filepath = self.filename_entry.get().strip()
        if not filepath.endswith(".xlsx"):
            filepath += ".xlsx"

        if not url or "ylilauta.org" not in url:
            messagebox.showerror("Error", "Please enter a valid Ylilauta thread URL.")
            return

        # Disable button while extraction
        self.get_btn.config(state="disabled")
        self.set_status("Starting...")

        try:
            posts = get_thread(url, status_callback=self.set_status)
            save_xlsx(posts, filepath)
            self.set_status(f"Done! Extracted {len(posts)} posts → {filepath}")
            messagebox.showinfo("Done", f"Saved {len(posts)} posts to:\n{filepath}")
        except ValueError as e:
            messagebox.showerror("Error", str(e))
            self.set_status(f"Error: {e}")
        except Exception as e:
            messagebox.showerror("Unexpected Error", f"{type(e).__name__}: {e}")
            self.set_status(f"Unexpected error: {e}")
        finally:
            # Re-enable button regardless of success/failure
            self.get_btn.config(state="normal")


if __name__ == "__main__":
    app = App()
    app.mainloop()
